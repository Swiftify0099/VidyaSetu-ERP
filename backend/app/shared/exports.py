"""
VidyaSetu ERP — PDF & Excel Export Service
==========================================
Generates PDFs and Excel files for:
  - Fee Receipts
  - Student ID Cards
  - Progress Reports / Mark Sheets
  - Attendance Certificates
  - Financial Ledger / Cash Book
  - Teacher Leave Reports

Uses: reportlab (PDF), openpyxl (Excel), Pillow (images)
"""
import io
import os
from datetime import date, datetime
from typing import Optional, Any
from decimal import Decimal

# ── PDF ───────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle,
    Spacer, HRFlowable, KeepTogether,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ── Excel ─────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from app.core.config import settings


# ── Color Palette ─────────────────────────────────────────────
PRIMARY   = colors.HexColor('#4f46e5')
PRIMARY_L = colors.HexColor('#ede9fe')
SUCCESS   = colors.HexColor('#059669')
DANGER    = colors.HexColor('#dc2626')
GRAY_100  = colors.HexColor('#f3f4f6')
GRAY_300  = colors.HexColor('#d1d5db')
GRAY_600  = colors.HexColor('#4b5563')
GRAY_900  = colors.HexColor('#111827')
WHITE     = colors.white

SCHOOL_NAME    = settings.SCHOOL_NAME
SCHOOL_ADDRESS = settings.SCHOOL_ADDRESS
SCHOOL_PHONE   = settings.SCHOOL_PHONE


def _header_table_style() -> TableStyle:
    return TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0), PRIMARY),
        ('TEXTCOLOR',   (0, 0), (-1, 0), WHITE),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0), 9),
        ('ALIGN',       (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING',  (0, 0), (-1, 0), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, GRAY_100]),
        ('FONTSIZE',    (0, 1), (-1, -1), 8),
        ('GRID',        (0, 0), (-1, -1), 0.4, GRAY_300),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',(0, 0), (-1, -1), 6),
    ])


def _school_header(styles: dict) -> list:
    """School header block reused across all PDFs."""
    elements = [
        Paragraph(SCHOOL_NAME, ParagraphStyle(
            'SchoolName', fontSize=16, fontName='Helvetica-Bold',
            alignment=TA_CENTER, textColor=PRIMARY,
        )),
        Paragraph(SCHOOL_ADDRESS or 'Vidya Nagar, Maharashtra', ParagraphStyle(
            'SchoolAddr', fontSize=9, alignment=TA_CENTER, textColor=GRAY_600,
        )),
        Paragraph(f"Phone: {SCHOOL_PHONE or 'N/A'}", ParagraphStyle(
            'SchoolPhone', fontSize=8, alignment=TA_CENTER, textColor=GRAY_600,
        )),
        HRFlowable(width="100%", thickness=1.5, color=PRIMARY, spaceAfter=6),
    ]
    return elements


# ═══════════════════════════════════════════════════════════════
# FEE RECEIPT PDF
# ═══════════════════════════════════════════════════════════════

def generate_fee_receipt_pdf(receipt: dict) -> bytes:
    """
    Generate a printable A4 fee receipt PDF.
    receipt dict keys:
      receipt_number, student_name, gr_number, standard, division,
      father_name, total_amount, payment_mode, receipt_date,
      collected_by_name, fee_items (list of {name, amount})
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=15*mm, leftMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )
    styles = getSampleStyleSheet()
    elements = []

    # Header
    elements.extend(_school_header(styles))

    # Receipt title
    elements.append(Paragraph("FEE RECEIPT", ParagraphStyle(
        'Title', fontSize=14, fontName='Helvetica-Bold',
        alignment=TA_CENTER, textColor=PRIMARY, spaceBefore=4, spaceAfter=8,
    )))

    # Receipt meta table
    meta_data = [
        ['Receipt No.', receipt.get('receipt_number', '—'), 'Date', receipt.get('receipt_date', '—')],
        ['GR Number',   receipt.get('gr_number', '—'),      'Academic Year', receipt.get('academic_year', settings.CURRENT_ACADEMIC_YEAR)],
    ]
    meta_table = Table(meta_data, colWidths=[35*mm, 65*mm, 35*mm, 45*mm])
    meta_table.setStyle(TableStyle([
        ('FONTNAME',    (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME',    (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, -1), 9),
        ('GRID',        (0, 0), (-1, -1), 0.4, GRAY_300),
        ('BACKGROUND',  (0, 0), (0, -1), GRAY_100),
        ('BACKGROUND',  (2, 0), (2, -1), GRAY_100),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',(0, 0), (-1, -1), 6),
        ('TOPPADDING',  (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 4*mm))

    # Student info
    student_data = [
        ['Student Name', receipt.get('student_name', '—')],
        ['Standard / Division', f"{receipt.get('standard', '')} / {receipt.get('division', '')}"],
        ["Father's Name", receipt.get('father_name', '—')],
    ]
    student_table = Table(student_data, colWidths=[55*mm, 125*mm])
    student_table.setStyle(TableStyle([
        ('FONTNAME',    (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, -1), 9),
        ('GRID',        (0, 0), (-1, -1), 0.4, GRAY_300),
        ('BACKGROUND',  (0, 0), (0, -1), PRIMARY_L),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING',  (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(student_table)
    elements.append(Spacer(1, 6*mm))

    # Fee breakdown
    fee_items = receipt.get('fee_items', [])
    fee_data = [['#', 'Fee Head', 'Amount (₹)']]
    total = Decimal('0')
    for i, item in enumerate(fee_items, 1):
        amt = Decimal(str(item.get('amount', 0)))
        total += amt
        fee_data.append([str(i), item.get('name', ''), f"₹ {amt:,.2f}"])
    fee_data.append(['', 'TOTAL', f"₹ {total:,.2f}"])

    fee_table = Table(fee_data, colWidths=[10*mm, 140*mm, 30*mm])
    ts = _header_table_style()
    ts.add('FONTNAME',  (0, -1), (-1, -1), 'Helvetica-Bold')
    ts.add('BACKGROUND',(0, -1), (-1, -1), PRIMARY)
    ts.add('TEXTCOLOR', (0, -1), (-1, -1), WHITE)
    ts.add('ALIGN',     (2, 0),  (2, -1),  'RIGHT')
    fee_table.setStyle(ts)
    elements.append(fee_table)
    elements.append(Spacer(1, 6*mm))

    # Payment info
    pay_data = [
        ['Payment Mode', receipt.get('payment_mode', '—').upper(), 'Collected By', receipt.get('collected_by_name', '—')],
    ]
    pay_table = Table(pay_data, colWidths=[35*mm, 65*mm, 35*mm, 45*mm])
    pay_table.setStyle(TableStyle([
        ('FONTNAME',    (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME',    (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, -1), 9),
        ('BACKGROUND',  (0, 0), (0, -1), GRAY_100),
        ('BACKGROUND',  (2, 0), (2, -1), GRAY_100),
        ('GRID',        (0, 0), (-1, -1), 0.4, GRAY_300),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING',  (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(pay_table)
    elements.append(Spacer(1, 15*mm))

    # Signatures row
    sig_data = [['Cashier Signature', '', 'Principal Seal & Signature']]
    sig_table = Table(sig_data, colWidths=[60*mm, 60*mm, 60*mm])
    sig_table.setStyle(TableStyle([
        ('ALIGN',     (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE',  (0, 0), (-1, -1), 8),
        ('FONTNAME',  (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, -1), GRAY_600),
        ('TOPPADDING',(0, 0), (-1, -1), 20),
        ('BOX',       (0, 0), (0, 0),  0.5, GRAY_600),
        ('BOX',       (2, 0), (2, 0),  0.5, GRAY_600),
    ]))
    elements.append(sig_table)

    # Footer
    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph(
        "This is a computer-generated receipt. No signature required.",
        ParagraphStyle('Footer', fontSize=7, alignment=TA_CENTER, textColor=GRAY_600),
    ))

    doc.build(elements)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════
# STUDENT MARK SHEET PDF
# ═══════════════════════════════════════════════════════════════

def generate_mark_sheet_pdf(data: dict) -> bytes:
    """
    Generate a student progress/mark sheet PDF.
    data keys:
      student (dict), exam_name, academic_year, subjects (list of {subject, max, obtained, grade})
      total_marks, max_marks, percentage, result, rank
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=15*mm, leftMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.extend(_school_header(styles))

    student = data.get('student', {})
    elements.append(Paragraph(
        f"MARK SHEET — {data.get('exam_name', 'Examination')}",
        ParagraphStyle('Title', fontSize=13, fontName='Helvetica-Bold',
                       alignment=TA_CENTER, textColor=PRIMARY, spaceBefore=4, spaceAfter=8),
    ))

    # Student info
    info_data = [
        ['Student Name', student.get('full_name', '—'), 'GR No.', student.get('gr_number', '—')],
        ['Standard',     f"{student.get('standard', '')} / {student.get('division', '')}",
         'Roll No.',     str(student.get('roll_number', '—'))],
        ['Academic Year', data.get('academic_year', settings.CURRENT_ACADEMIC_YEAR),
         'Exam Date',    data.get('exam_date', '—')],
    ]
    info_table = Table(info_data, colWidths=[35*mm, 65*mm, 25*mm, 55*mm])
    info_table.setStyle(TableStyle([
        ('FONTNAME',    (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME',    (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, -1), 9),
        ('GRID',        (0, 0), (-1, -1), 0.4, GRAY_300),
        ('BACKGROUND',  (0, 0), (0, -1), PRIMARY_L),
        ('BACKGROUND',  (2, 0), (2, -1), PRIMARY_L),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING',  (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 6*mm))

    # Subject-wise marks
    subjects = data.get('subjects', [])
    marks_data = [['Subject', 'Max Marks', 'Obtained', 'Grade', 'Remarks']]
    for s in subjects:
        obtained = s.get('obtained', 0)
        max_m = s.get('max', 0)
        passing = s.get('passing', 35)
        failed = obtained < passing
        marks_data.append([
            s.get('subject', '—'),
            str(max_m),
            str(obtained),
            s.get('grade', '—'),
            'FAIL' if failed else 'PASS',
        ])

    marks_table = Table(marks_data, colWidths=[70*mm, 25*mm, 25*mm, 25*mm, 35*mm])
    ts = _header_table_style()
    marks_table.setStyle(ts)
    elements.append(marks_table)
    elements.append(Spacer(1, 4*mm))

    # Summary row
    percentage = data.get('percentage', 0)
    result = data.get('result', 'pending').upper()
    result_color = SUCCESS if result == 'PASS' else DANGER

    summary_data = [
        ['Total Marks', f"{data.get('total_marks', 0)} / {data.get('max_marks', 0)}",
         'Percentage',  f"{percentage:.1f}%",
         'Class Rank',  str(data.get('rank', '—'))],
        ['Result', result, 'Grade',
         data.get('overall_grade', '—'), '', ''],
    ]
    summary_table = Table(summary_data, colWidths=[30*mm, 40*mm, 30*mm, 30*mm, 20*mm, 30*mm])
    summary_table.setStyle(TableStyle([
        ('FONTNAME',    (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, -1), 9),
        ('GRID',        (0, 0), (-1, -1), 0.4, GRAY_300),
        ('BACKGROUND',  (0, 0), (0, -1), PRIMARY),
        ('TEXTCOLOR',   (0, 0), (0, -1), WHITE),
        ('BACKGROUND',  (2, 0), (2, -1), PRIMARY),
        ('TEXTCOLOR',   (2, 0), (2, -1), WHITE),
        ('BACKGROUND',  (4, 0), (4, -1), PRIMARY),
        ('TEXTCOLOR',   (4, 0), (4, -1), WHITE),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING',  (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 15*mm))

    # Signatures
    sig_data = [["Class Teacher's Signature", '', "Principal's Signature"]]
    sig_table = Table(sig_data, colWidths=[70*mm, 40*mm, 70*mm])
    sig_table.setStyle(TableStyle([
        ('ALIGN',    (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 20),
        ('BOX',      (0, 0), (0, 0),  0.5, GRAY_600),
        ('BOX',      (2, 0), (2, 0),  0.5, GRAY_600),
    ]))
    elements.append(sig_table)

    doc.build(elements)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════
# ATTENDANCE REPORT PDF
# ═══════════════════════════════════════════════════════════════

def generate_attendance_report_pdf(data: dict) -> bytes:
    """Attendance report for a class or student."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=15*mm, leftMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )
    elements = []
    styles = getSampleStyleSheet()
    elements.extend(_school_header(styles))

    elements.append(Paragraph(
        f"ATTENDANCE REPORT — {data.get('period', '')}",
        ParagraphStyle('Title', fontSize=13, fontName='Helvetica-Bold',
                       alignment=TA_CENTER, textColor=PRIMARY, spaceBefore=4, spaceAfter=8),
    ))

    elements.append(Paragraph(
        f"Standard: {data.get('standard', '')} | Division: {data.get('division', '')} | "
        f"Academic Year: {data.get('academic_year', settings.CURRENT_ACADEMIC_YEAR)}",
        ParagraphStyle('Sub', fontSize=9, alignment=TA_CENTER,
                       textColor=GRAY_600, spaceAfter=8),
    ))

    rows = data.get('rows', [])
    table_data = [['#', 'GR No.', 'Student Name', 'Total Days', 'Present', 'Absent', 'Leave', '%']]
    for i, row in enumerate(rows, 1):
        table_data.append([
            str(i), row.get('gr_number', '—'), row.get('student_name', '—'),
            str(row.get('total_days', 0)), str(row.get('present', 0)),
            str(row.get('absent', 0)), str(row.get('leave', 0)),
            f"{row.get('percentage', 0):.1f}%",
        ])

    t = Table(table_data, colWidths=[10*mm, 25*mm, 70*mm, 22*mm, 22*mm, 22*mm, 22*mm, 17*mm])
    t.setStyle(_header_table_style())
    elements.append(t)

    doc.build(elements)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════
# EXCEL EXPORTS
# ═══════════════════════════════════════════════════════════════

def _excel_header_style() -> dict:
    return {
        'fill': PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid'),
        'font': Font(bold=True, color='FFFFFF', size=10),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),   right=Side(style='thin'),
            top=Side(style='thin'),    bottom=Side(style='thin'),
        ),
    }


def _apply_header_style(ws, row: int, max_col: int):
    header_style = _excel_header_style()
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_style['fill']
        cell.font = header_style['font']
        cell.alignment = header_style['alignment']
        cell.border = header_style['border']


def generate_students_excel(students: list[dict]) -> bytes:
    """Export student list to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.sheet_properties.tabColor = "4F46E5"

    # Title
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = f"{SCHOOL_NAME} — Student List"
    title_cell.font = Font(bold=True, size=14, color='4F46E5')
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 24

    # Sub-title
    ws.merge_cells('A2:L2')
    ws['A2'].value = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=9, color='6B7280')

    # Headers
    headers = [
        'GR Number', 'Admission No.', 'Full Name', 'Standard', 'Division',
        'Roll No.', 'Gender', 'Date of Birth', 'Mobile', "Father's Name",
        'Category', 'Status',
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _apply_header_style(ws, 4, len(headers))
    ws.row_dimensions[4].height = 20

    # Data rows
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    for row_idx, student in enumerate(students, 5):
        row_data = [
            student.get('gr_number', ''), student.get('admission_number', ''),
            student.get('full_name', ''), student.get('standard', ''),
            student.get('division', ''), student.get('roll_number', ''),
            student.get('gender', ''), student.get('dob', ''),
            student.get('mobile', ''), student.get('father_name', ''),
            student.get('category', ''), student.get('status', ''),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    # Column widths
    col_widths = [15, 16, 30, 10, 10, 10, 10, 14, 14, 28, 12, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}4'

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_fee_collection_excel(receipts: list[dict], filters: dict) -> bytes:
    """Export fee collection report to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fee Collection"

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'].value = f"{SCHOOL_NAME} — Fee Collection Report"
    ws['A1'].font = Font(bold=True, size=13, color='4F46E5')
    ws['A1'].alignment = Alignment(horizontal='center')

    period = f"{filters.get('date_from', '')} to {filters.get('date_to', '')}"
    ws.merge_cells('A2:J2')
    ws['A2'].value = f"Period: {period} | Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=9, color='6B7280')

    headers = [
        'Receipt No.', 'Date', 'GR No.', 'Student Name', 'Standard',
        'Division', 'Amount (₹)', 'Payment Mode', 'Collected By', 'Status',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    _apply_header_style(ws, 4, len(headers))

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    total = Decimal('0')
    for row_idx, receipt in enumerate(receipts, 5):
        amount = Decimal(str(receipt.get('total_amount', 0)))
        total += amount
        row_data = [
            receipt.get('receipt_number', ''), receipt.get('receipt_date', ''),
            receipt.get('gr_number', ''), receipt.get('student_name', ''),
            receipt.get('standard', ''), receipt.get('division', ''),
            float(amount), receipt.get('payment_mode', '').upper(),
            receipt.get('collected_by_name', ''), receipt.get('status', ''),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if col_idx == 7:  # Amount column
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    # Total row
    total_row = len(receipts) + 5
    ws.cell(row=total_row, column=6, value='TOTAL').font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=7, value=float(total))
    total_cell.font = Font(bold=True, color='059669')
    total_cell.number_format = '₹#,##0.00'
    total_cell.alignment = Alignment(horizontal='right')

    col_widths = [18, 14, 12, 28, 10, 10, 15, 16, 22, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}4'

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_marks_excel(exam_name: str, standard: str, results: list[dict]) -> bytes:
    """Export exam mark sheet to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mark Sheet"

    ws.merge_cells('A1:K1')
    ws['A1'].value = f"{SCHOOL_NAME} — {exam_name} — Std {standard}"
    ws['A1'].font = Font(bold=True, size=13, color='4F46E5')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'Rank', 'GR No.', 'Student Name', 'Roll No.',
        'Total Marks', 'Max Marks', 'Percentage', 'Grade', 'Result',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _apply_header_style(ws, 3, len(headers))

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    for row_idx, result in enumerate(results, 4):
        row_data = [
            result.get('rank', '—'), result.get('gr_number', ''),
            result.get('student_name', ''), result.get('roll_number', ''),
            result.get('total_marks', 0), result.get('max_marks', 0),
            result.get('percentage', 0), result.get('grade', '—'),
            result.get('result', '').upper(),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 7:
                cell.number_format = '0.00"%"'
            # Red for FAIL
            if col_idx == 9 and value == 'FAIL':
                cell.font = Font(bold=True, color='DC2626')
            elif col_idx == 9:
                cell.font = Font(bold=True, color='059669')
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    col_widths = [8, 14, 30, 10, 14, 12, 14, 10, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A4'
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
